import openpyxl
import os 
import time
#Opens excel file from directory and sets workbook object and sheet object to active sheet (first sheet)
path = "AdamFYP2TrainingData.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row
col = sheet_obj.max_column

#Function that collects and combine values from Kemahiran Hidup, Asas Sains Komputer, Reka Bentuk Teknologi column
def updateCollatField(targetCol,firstCol):
    #Sets possible range of grades to compare values to
    grade = ["A","B","C","D","E","Gagal"]
    for i in range(2, row + 1):
        #initializes cell objects to current row of iteration
        cell_obj_KH = sheet_obj.cell(row=i,column=firstCol)
        cell_obj_ASK = sheet_obj.cell(row=i,column=firstCol+1)
        cell_obj_RBT = sheet_obj.cell(row=i,column=firstCol+2)
        #elif statements to check if cell is empty or not. If cell has a valid grade, assign that grade to the target column
        if cell_obj_KH.value in grade:
            cell_obj_temp = sheet_obj.cell(row=i,column=targetCol)
            cell_obj_temp.value=cell_obj_KH.value
        elif cell_obj_ASK.value in grade:
            cell_obj_temp = sheet_obj.cell(row=i,column=targetCol)
            cell_obj_temp.value=cell_obj_ASK.value
        elif cell_obj_RBT.value in grade:
            cell_obj_temp = sheet_obj.cell(row=i,column=targetCol)
            cell_obj_temp.value=cell_obj_RBT.value
        else:
            cell_obj_temp = sheet_obj.cell(row=i,column=targetCol)
            cell_obj_temp.value = "N/A"

#Function that appends new column at the end
def addNewColumn(colName):
    print("{} will be added at {}".format(colName,col+1))
    sheet_obj.insert_cols(col+1)
    cell_obj = sheet_obj.cell(row=1,column=col+1)
    cell_obj.value = colName

def imputeCol(targetCol):
    gradeValueDict ={"A":1,"B":2,"C":3,"D":4,"E":5,"Gagal":6}
    valueGradeDict = {1:"A",2:"B",3:"C",4:"D",5:"E",6:"F"}
    sum=0
    grade = ["A","B","C","D","E","Gagal"]
    filledRows = row
    for i in range(2,row+1):
        cell_obj=sheet_obj.cell(row=i,column=targetCol)
        if cell_obj.value in grade:
            sum += gradeValueDict[cell_obj.value]
        else:
            filledRows-=1
    average = sum/filledRows
    average = round(average,0)
    average = valueGradeDict[average]
    for i in range(2,row+1):
        cell_obj = sheet_obj.cell(row = i,column=targetCol)
        if cell_obj.value not in grade:
            cell_obj.value = average
def getColNumber(colName):
    for i in range(2,col+1):
        col_obj = sheet_obj.cell(row=1,column=i)
        if col_obj.value == colName:
            return i
    return "Column does not exist"

#Gets number of rows filled
print("Number of rows is {}".format(row))
print("[1] Create new column")
print("[2] Collate columns")
print("[3] Data imputation")
choice = input("Which operation would you like to do?: ")
if choice == "1":
    colName = input("What would you like the column name to be?: ")
    colResponse = getColNumber(colName)
    if colResponse == "Column does not exist":
        addNewColumn(colName)
        print("New column added")
    else:
        print("Column name already exist at column: {}".format(colResponse))
elif choice == "2":
    colName = input("What would you like the starting column name to be?: ")
    colResponse = getColNumber(colName)
    firstCol = colResponse
    if firstCol == "Column does not exist":
        print("Im sorry but we couldn't find that column, invalid operation")
    else:
        colName = input("What would you like the output column name to be?: ")
        colResponse = getColNumber(colName)
        if firstCol == "Column does not exist":
            print("Im sorry but we couldn't find that column, invalid operation")
        else:
            updateCollatField(colResponse,firstCol) 
elif choice == "3":
    colName = input("Which column would you want to impute?: ")
    colResponse = getColNumber(colName)
    if colResponse == "Column does not exist":
        print("Im sorry but we couldn't find that column, invalid operation")
    else:
        print("HELLO HELLO colResponse is {}".format(colResponse))
        imputeCol(colResponse)




wb_obj.save("AdamFYP2TrainingData.xlsx")
print("Successfully cleaned")   